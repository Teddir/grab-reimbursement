from openpyxl import load_workbook
import re

def inspect_template(path):
    print(f"Inspecting: {path}")
    wb = load_workbook(path)
    ws = wb.active
    print(f"Active Sheet: {ws.title}")
    
    placeholder_pattern = re.compile(r'value_[a-zA-Z0-9_]+')
    found_placeholders = []
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                matches = placeholder_pattern.findall(cell.value)
                for match in matches:
                    found_placeholders.append((cell.coordinate, match))
    
    if found_placeholders:
        print("\nFound Placeholders:")
        for coord, name in found_placeholders:
            print(f"- {coord}: {name}")
    else:
        print("\nNo placeholders found using 'value_*' pattern.")

    # Check for merged cells
    if ws.merged_cells:
        print(f"\nMerged Cells: {len(ws.merged_cells.ranges)} ranges found.")

    # Check for row/column dimensions
    print(f"\nMax Row: {ws.max_row}")
    print(f"Max Column: {ws.max_column}")

if __name__ == "__main__":
    inspect_template("/Users/mac/Documents/asiasistem/grab-reimbursement/form_template.xlsx")
