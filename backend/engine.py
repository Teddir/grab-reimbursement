import re
import copy
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from io import BytesIO

class ExcelTemplateEngine:
    def __init__(self, template_path):
        self.template_path = template_path
        self.wb = load_workbook(template_path)
        self.placeholder_pattern = re.compile(r'value_[a-zA-Z0-9_]+')

    def find_placeholders(self, ws):
        placeholders = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    matches = self.placeholder_pattern.findall(cell.value)
                    for match in matches:
                        placeholders.append({
                            'name': match,
                            'cell': cell,
                            'coord': cell.coordinate,
                            'row': cell.row,
                            'col': cell.column
                        })
        return placeholders

    def copy_row_style(self, source_ws, source_row_idx, target_ws, target_row_idx):
        """
        Deep copies styles and dimensions from source row to target row.
        """
        # Copy row height
        target_ws.row_dimensions[target_row_idx].height = source_ws.row_dimensions[source_row_idx].height
        
        for col_idx in range(1, source_ws.max_column + 1):
            source_cell = source_ws.cell(row=source_row_idx, column=col_idx)
            target_cell = target_ws.cell(row=target_row_idx, column=col_idx)
            
            if source_cell.has_style:
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = copy.copy(source_cell.number_format)
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.alignment = copy.copy(source_cell.alignment)
            
            target_cell.value = source_cell.value

    def duplicate_template_row(self, ws, template_row_idx, count):
        """
        Duplicates a template row multiple times, preserving styles and merged cells.
        Useful for dynamic lists of receipts.
        """
        if count <= 1:
            return

        # Insert rows after the template row
        ws.insert_rows(template_row_idx + 1, amount=count - 1)
        
        # Track merged cells that start in the template row
        merged_ranges_to_clone = []
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == template_row_idx and merged_range.max_row == template_row_idx:
                merged_ranges_to_clone.append(merged_range)

        for i in range(1, count):
            target_row_idx = template_row_idx + i
            self.copy_row_style(ws, template_row_idx, ws, target_row_idx)
            
            # Re-apply merged cells for the new row
            for m_range in merged_ranges_to_clone:
                new_range = f"{get_column_letter(m_range.min_col)}{target_row_idx}:{get_column_letter(m_range.max_col)}{target_row_idx}"
                ws.merge_cells(new_range)

    def fill_template(self, data_list, output_path):
        """
        data_list: List of dicts, each dict contains key-value pairs for placeholders.
        """
        ws = self.wb.active
        print(f"Filling template on sheet: {ws.title}")
        
        # 1. Detect which row contains the placeholders for the repeatable data
        all_placeholders = self.find_placeholders(ws)
        if not all_placeholders:
            print("Warning: No placeholders found in the active worksheet.")
            self.wb.save(output_path)
            return

        # Find rows that contain 'value_no' or similar repeatable markers
        repeat_row_idx = None
        repeatable_markers = ['value_no', 'value_nomor_order_grab', 'value_tanggal_perjalanan']
        for p in all_placeholders:
            if p['name'] in repeatable_markers: # Exact match to avoid partial matches like value_no_dok
                repeat_row_idx = p['row']
                break
        
        if not repeat_row_idx:
            repeat_row_idx = all_placeholders[0]['row']
            print(f"Assuming repeatable row is {repeat_row_idx}")

        # 2. Duplicate the row if we have multiple items
        if repeat_row_idx and len(data_list) > 1:
            num_new_rows = len(data_list) - 1
            print(f"Duplicating row {repeat_row_idx} into {num_new_rows} new rows")
            
            # Identify merged cells in the template row before inserting
            template_merges = []
            for m_range in list(ws.merged_cells.ranges):
                if m_range.min_row <= repeat_row_idx <= m_range.max_row:
                    template_merges.append(m_range)
            
            # Insert rows - this shifts everything below
            ws.insert_rows(repeat_row_idx + 1, amount=num_new_rows)
            
            # Capture the template row values and styles
            # We must iterate over all columns to ensure borders are captured for empty cells too
            max_col = ws.max_column
            source_cells = [ws.cell(row=repeat_row_idx, column=c) for c in range(1, max_col + 1)]

            for i in range(1, len(data_list)):
                target_row_idx = repeat_row_idx + i
                ws.row_dimensions[target_row_idx].height = ws.row_dimensions[repeat_row_idx].height
                
                # Surgical cleanup: only remove merges that were created/shifted into this specific new row
                # but belong to the item table structure
                for m_range in list(ws.merged_cells.ranges):
                    if m_range.min_row == target_row_idx and m_range.max_row == target_row_idx:
                        try:
                            ws.unmerge_cells(m_range.coord)
                        except:
                            pass

                # Deep copy values and styles for every cell in the row
                for col_idx, source_cell in enumerate(source_cells, 1):
                    target_cell = ws.cell(row=target_row_idx, column=col_idx)
                    target_cell.value = source_cell.value
                    if source_cell.has_style:
                        target_cell.font = copy.copy(source_cell.font)
                        target_cell.border = copy.copy(source_cell.border)
                        target_cell.fill = copy.copy(source_cell.fill)
                        target_cell.number_format = copy.copy(source_cell.number_format)
                        target_cell.alignment = copy.copy(source_cell.alignment)

                # Re-apply merged cells for the new row exactly as they are in the template
                for m_range in template_merges:
                    ws.merge_cells(
                        start_row=target_row_idx,
                        start_column=m_range.min_col,
                        end_row=target_row_idx,
                        end_column=m_range.max_col
                    )

        # 3. Substitute values in the table rows
        for i, data in enumerate(data_list):
            current_row_idx = repeat_row_idx + i if repeat_row_idx else i + 1
            print(f"--- Filling Excel Row {current_row_idx} ---")
            
            # Use iter_rows for stability
            row_cells = list(ws.iter_rows(min_row=current_row_idx, max_row=current_row_idx, min_col=1, max_col=ws.max_column))[0]
            for cell in row_cells:
                if cell.value and isinstance(cell.value, str):
                    new_value = str(cell.value)
                    original_val = new_value
                    
                    sorted_keys = sorted(data.keys(), key=len, reverse=True)
                    for key in sorted_keys:
                        if key in new_value:
                            val = data[key]
                            if key == 'value_image_receipt' and val:
                                self.insert_image(ws, val, cell.coordinate)
                                new_value = new_value.replace(key, "") 
                                print(f"  Inserted image at {cell.coordinate}")
                            else:
                                new_value = new_value.replace(key, str(val))
                                print(f"  Replaced {key} with '{val}' at {cell.coordinate}")
                    
                    if new_value != original_val:
                        cell.value = new_value

        # 4. Global Pass: Fill any remaining placeholders (header/footer)
        global_data = {}
        if data_list:
            # Merge all unique keys from all data items
            for d in reversed(data_list):
                global_data.update(d)
        
        # Calculate total
        try:
            def clean_float(val):
                if not val: return 0.0
                s = str(val).replace('IDR', '').replace('RP', '').replace(' ', '').replace(',', '')
                if '.' in s: s = s.replace('.', '')
                return float(s) if s else 0.0

            total_cost = sum(clean_float(d.get("value_total_biaya", 0)) for d in data_list)
            global_data["value_total_biaya"] = f"Rp {total_cost:,.0f}"
            global_data["value_total_fare"] = f"Rp {total_cost:,.0f}"
        except Exception as e:
            print(f"Total calculation error: {e}")
        
        # Re-scan the entire worksheet for remaining placeholders
        # We also re-apply signature/footer merges that might have been shifted or corrupted
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    new_value = str(cell.value)
                    replaced = False
                    
                    # Search for placeholders
                    matches = self.placeholder_pattern.findall(new_value)
                    if matches:
                        sorted_matches = sorted(matches, key=len, reverse=True)
                        for match in sorted_matches:
                            if match in global_data:
                                val = global_data[match]
                                if match == 'value_image_receipt' and val:
                                    self.insert_image(ws, val, cell.coordinate)
                                    new_value = new_value.replace(match, "")
                                else:
                                    new_value = new_value.replace(match, str(val))
                                replaced = True
                        
                        if replaced:
                            cell.value = new_value
        
        # 5. Final Footer Correction: Force apply user-specified signature/total merges
        # This ensures the footer layout is exactly as requested even after row shifts
        last_data_row = repeat_row_idx + len(data_list)
        for row_idx in range(last_data_row, ws.max_row + 1):
            for cell in ws[row_idx]:
                if not cell.value: continue
                val = str(cell.value)
                
                # Total Biaya (Label and Value)
                if "Total Biaya" in val or "Rp" in val:
                    # Search nearby for labels or values to identify the row
                    # Apply: Total Biaya (B-C), Value (D-H)
                    try:
                        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=3)
                        ws.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx, end_column=8)
                    except: pass
                
                # Pemohon / Signature block 1
                if "Pemohon" in val or any(x in val for x in [global_data.get("value_nama_karyawan", ""), global_data.get("value_pemohon", "")] if x):
                    # Apply: B-F merge
                    try:
                        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
                    except: pass
                
                # HR/GA / Signature block 2
                if "HR/GA" in val or global_data.get("value_hr", "") in val:
                    # Apply: J-L merge
                    try:
                        ws.merge_cells(start_row=row_idx, start_column=10, end_row=row_idx, end_column=12)
                    except: pass

        self.wb.save(output_path)
        print(f"Successfully saved filled report to: {output_path}")

    def insert_image(self, ws, img_data, coordinate):
        """
        Inserts an image into the specified cell, attempting to fit it within the cell boundaries.
        """
        img = OpenpyxlImage(BytesIO(img_data))
        
        # Scale image to a reasonable size while maintaining aspect ratio
        # Standard Excel cell is roughly 64x20 pixels at 100% zoom
        # We'll aim for about 120px height and proportional width
        max_height = 120
        scale = max_height / img.height
        img.width = int(img.width * scale)
        img.height = max_height
        
        # Anchor the image to the top-left of the cell
        img.anchor = coordinate
        ws.add_image(img)

    def save(self, path):
        self.wb.save(path)
